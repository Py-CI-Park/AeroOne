"""Bounded CSV/XLSX/JSON loading for the chart studio.

Every format is validated before a pandas ``DataFrame`` is constructed.  CSV
and JSON are decoded incrementally, while XLSX/XLSM archives are preflighted
and streamed with openpyxl's read-only mode.
"""

from __future__ import annotations

import codecs
import json
import math
import re
import zipfile
import zlib
from dataclasses import dataclass
from importlib.util import find_spec
from io import BytesIO, StringIO
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree.ElementTree import ParseError, XMLPullParser

import numpy as np
import pandas as pd
from app.modules.office_tools.limits import (
    MAX_ZIP_CENTRAL_DIRECTORY_BYTES,
    OFFICE_STREAM_CHUNK_BYTES,
    UploadSizeLimitExceeded,
    preflight_zip_central_directory,
)

# 지원 확장자. xlsx/xlsm 은 openpyxl 설치 여부에 따라 런타임에 다시 확인한다.
SUPPORTED_DATA_SUFFIXES = {'.csv', '.xlsx', '.xlsm', '.json'}
_EXCEL_SUFFIXES = {'.xlsx', '.xlsm'}
_CSV_ENCODINGS = ('utf-8-sig', 'utf-8', 'cp949', 'euc-kr')

# Input caps are deliberately independent from pandas so malformed content is
# rejected before its broad, allocation-heavy parsers construct a DataFrame.
MAX_CHART_DATA_COLUMNS = 128
MAX_CHART_DATA_CELL_CHARS = 64 * 1024
MAX_CHART_DECODED_CHARS = 20 * 1024 * 1024
MAX_CHART_DATA_CELLS = 1_000_000
MAX_CHART_XLSX_ENTRIES = 1_000
MAX_CHART_XLSX_EXPANDED_BYTES = 100 * 1024 * 1024
MAX_CHART_XLSX_MEMBER_BYTES = 20 * 1024 * 1024
MAX_CHART_XLSX_COMPRESSION_RATIO = 100
_DECODE_CHUNK_BYTES = OFFICE_STREAM_CHUNK_BYTES
_JSON_TOKEN_CHUNK_CHARS = 8 * 1024
MAX_CHART_JSON_OBJECT_CHARS = MAX_CHART_DATA_COLUMNS * (MAX_CHART_DATA_CELL_CHARS + 1024)
MAX_CHART_JSON_SCALAR_TOKEN_CHARS = 1024
MAX_CHART_XLSX_COMPRESSED_BYTES = 20 * 1024 * 1024
MAX_CHART_XLSX_FILENAME_BYTES = 1024
MAX_CHART_XLSX_CENTRAL_DIRECTORY_BYTES = MAX_ZIP_CENTRAL_DIRECTORY_BYTES
MAX_CHART_XLSX_METADATA_BYTES = 2 * 1024 * 1024
MAX_CHART_XLSX_SHARED_STRINGS = 250_000
MAX_CHART_XLSX_STYLE_RECORDS = 10_000
MAX_CHART_XLSX_NUMBER_FORMATS = 10_000
MAX_CHART_XLSX_FONTS = 10_000
MAX_CHART_XLSX_FILLS = 10_000
MAX_CHART_XLSX_BORDERS = 10_000
MAX_CHART_XLSX_ALIGNMENTS = 10_000
MAX_CHART_XLSX_PROTECTIONS = 10_000
MAX_CHART_XLSX_NAMED_STYLES = 10_000
MAX_CHART_XLSX_DIFFERENTIAL_STYLES = 10_000
MAX_CHART_XLSX_GRADIENT_STOPS = 10_000
MAX_CHART_XLSX_TABLE_STYLE_ELEMENTS = 10_000
MAX_CHART_XLSX_INDEXED_COLORS = 10_000
MAX_CHART_XLSX_MRU_COLORS = 10_000
MAX_CHART_XLSX_RELATIONSHIPS = 2_000
MAX_CHART_XLSX_WORKBOOK_RECORDS = 10_000
MAX_CHART_XLSX_WORKBOOK_SHEETS = 1_000
MAX_CHART_XLSX_DEFINED_NAMES = 10_000
MAX_CHART_XLSX_EXTERNAL_REFERENCES = 1_000
MAX_CHART_XLSX_PIVOT_CACHES = 1_000
MAX_CHART_XLSX_WORKBOOK_VIEWS = 128
MAX_CHART_XLSX_FUNCTION_GROUPS = 1_000
MAX_CHART_XLSX_CUSTOM_WORKBOOK_VIEWS = 1_000
MAX_CHART_XLSX_FILE_RECOVERY_RECORDS = 1_000
MAX_CHART_XLSX_SMART_TAG_TYPES = 1_000
MAX_CHART_XLSX_WEB_PUBLISH_OBJECTS = 1_000
MAX_CHART_XLSX_INLINE_PAYLOADS_PER_CELL = 128
_OOXML_SPREADSHEET_NAMESPACE = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_OOXML_PACKAGE_RELATIONSHIPS_NAMESPACE = 'http://schemas.openxmlformats.org/package/2006/relationships'
_OOXML_DOCUMENT_RELATIONSHIPS_NAMESPACE = (
    'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
)
_OOXML_RELATIONSHIP_ID_ATTRIBUTE = f'{{{_OOXML_DOCUMENT_RELATIONSHIPS_NAMESPACE}}}id'
_OOXML_WORKSHEET_RELATIONSHIP_TYPE = (
    f'{_OOXML_DOCUMENT_RELATIONSHIPS_NAMESPACE}/worksheet'
)
MAX_CHART_XLSX_CONTENT_TYPES = 2_000
_XLSX_CELL_REFERENCE_RE = re.compile(r'^([A-Z]{1,3})([1-9][0-9]*)$')
_XLSX_STYLE_ALLOCATION_TAGS = frozenset(
    {
        'numFmt',
        'font',
        'fill',
        'border',
        'alignment',
        'protection',
        'xf',
        'cellStyle',
        'dxf',
        'tableStyle',
        'tableStyleElement',
        'stop',
        'color',
        'rgbColor',
        'indexedColor',
        'mruColor',
    }
)
_XLSX_WORKBOOK_ALLOCATION_TAGS = frozenset(
    {
        'workbookView',
        'sheet',
        'functionGroup',
        'externalReference',
        'definedName',
        'customWorkbookView',
        'pivotCache',
        'smartTagType',
        'fileRecoveryPr',
        'webPublishObject',
    }
)


def _unique_columns(columns) -> list[str]:
    """중복/빈 열 이름을 안정적으로 유일화한다(``col`` → ``col``, ``col_2`` …)."""
    next_suffix: dict[str, int] = {}
    reserved: set[str] = set()
    result: list[str] = []
    for raw in columns:
        base = str(raw).strip() or 'column'
        suffix = next_suffix.get(base, 1)
        name = base if suffix == 1 else f'{base}_{suffix}'
        while name in reserved:
            suffix += 1
            name = f'{base}_{suffix}'
        next_suffix[base] = suffix + 1
        reserved.add(name)
        result.append(name)
    return result


def _raise_limit(label: str, limit: int) -> None:
    raise ValueError(f'{label}이(가) 상한 {limit:,}을(를) 초과했습니다')


def _require_rectangular_cell_budget(row_count: int, column_count: int) -> None:
    """Reject a final table shape before rows or a DataFrame can expand to it."""
    if row_count < 0 or column_count < 0:
        raise ValueError('데이터 표 크기가 올바르지 않습니다')
    if column_count and row_count > MAX_CHART_DATA_CELLS // column_count:
        _raise_limit('데이터 셀 수', MAX_CHART_DATA_CELLS)


def _dataframe_from_rows(header: list[Any], rows: list[list[Any]]) -> pd.DataFrame:
    """Validated records only을 DataFrame으로 materialize하고 숫자 열을 복구한다."""
    if not rows:
        raise ValueError('데이터 파일에 행이 없습니다')
    if len(header) > MAX_CHART_DATA_COLUMNS:
        _raise_limit('데이터 열 수', MAX_CHART_DATA_COLUMNS)
    _require_rectangular_cell_budget(len(rows), len(header))
    if any(len(row) > len(header) for row in rows):
        raise ValueError('데이터 행의 열 수가 헤더보다 많습니다')

    frame = pd.DataFrame(rows, columns=_unique_columns(header))
    frame = frame.replace('', np.nan)
    for name in frame.columns:
        series = frame[name]
        non_null = series.notna()
        if not bool(non_null.any()):
            continue
        numeric = pd.to_numeric(series, errors='coerce')
        if bool(numeric[non_null].notna().all()):
            frame[name] = numeric
    return frame


def _read_csv(data: bytes, max_rows: int) -> pd.DataFrame:
    last_error: UnicodeDecodeError | None = None
    for encoding in _CSV_ENCODINGS:
        try:
            header, rows = _parse_csv_rows(data, encoding=encoding, max_rows=max_rows)
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        return _dataframe_from_rows(header, rows)
    raise ValueError(f'CSV 를 해독하지 못했습니다: {last_error}')


def _parse_csv_rows(data: bytes, *, encoding: str, max_rows: int) -> tuple[list[str], list[list[str]]]:
    """CSV 문자를 청크로 해석해 행/열/셀/총 해독량 상한을 먼저 적용한다."""
    decoder = codecs.getincrementaldecoder(encoding)(errors='strict')
    records: list[list[str]] = []
    row: list[str] = []
    cell: list[str] = []
    cell_chars = 0
    decoded_chars = 0
    in_quotes = False
    quote_pending = False
    row_has_content = False
    skip_line_feed = False

    def begin_data_row() -> None:
        if not records or row_has_content or row or cell:
            return
        header_width = len(records[0])
        data_rows = len(records)
        if data_rows > max_rows:
            _raise_limit('데이터 행 수', max_rows)
        _require_rectangular_cell_budget(data_rows, header_width)

    def append_cell_character(character: str) -> None:
        nonlocal cell_chars, row_has_content
        begin_data_row()
        cell_chars += 1
        if cell_chars > MAX_CHART_DATA_CELL_CHARS:
            _raise_limit('CSV 셀 문자 수', MAX_CHART_DATA_CELL_CHARS)
        cell.append(character)
        row_has_content = True

    def finish_cell() -> None:
        nonlocal cell, cell_chars
        row.append(''.join(cell))
        if len(row) > MAX_CHART_DATA_COLUMNS:
            _raise_limit('데이터 열 수', MAX_CHART_DATA_COLUMNS)
        cell = []
        cell_chars = 0

    def finish_row() -> None:
        nonlocal row, row_has_content
        if not row_has_content and not row and not cell:
            return
        finish_cell()
        if not records:
            records.append(row)
        else:
            header_width = len(records[0])
            if len(row) > header_width:
                raise ValueError('데이터 행의 열 수가 헤더보다 많습니다')
            data_rows = len(records)
            if data_rows > max_rows:
                _raise_limit('데이터 행 수', max_rows)
            _require_rectangular_cell_budget(data_rows, header_width)
            records.append(row)
        row = []
        row_has_content = False

    def consume(character: str) -> None:
        nonlocal in_quotes, quote_pending, row_has_content, skip_line_feed
        if skip_line_feed:
            skip_line_feed = False
            if character == '\n':
                return
        if character == '\r':
            skip_line_feed = True
            character = '\n'

        if in_quotes:
            if quote_pending:
                if character == '"':
                    append_cell_character('"')
                    quote_pending = False
                    return
                in_quotes = False
                quote_pending = False
                consume(character)
                return
            if character == '"':
                quote_pending = True
            else:
                append_cell_character(character)
            return

        if character == '"':
            if cell:
                raise ValueError('CSV 따옴표 위치가 올바르지 않습니다')
            in_quotes = True
            begin_data_row()
            row_has_content = True
        elif character == ',':
            begin_data_row()
            row_has_content = True
            finish_cell()
        elif character == '\n':
            finish_row()
        else:
            append_cell_character(character)

    for offset in range(0, len(data), _DECODE_CHUNK_BYTES):
        decoded = decoder.decode(data[offset:offset + _DECODE_CHUNK_BYTES], final=False)
        decoded_chars += len(decoded)
        if decoded_chars > MAX_CHART_DECODED_CHARS:
            _raise_limit('CSV 해독 문자 수', MAX_CHART_DECODED_CHARS)
        for character in decoded:
            consume(character)
    decoded = decoder.decode(b'', final=True)
    decoded_chars += len(decoded)
    if decoded_chars > MAX_CHART_DECODED_CHARS:
        _raise_limit('CSV 해독 문자 수', MAX_CHART_DECODED_CHARS)
    for character in decoded:
        consume(character)

    if in_quotes:
        if quote_pending:
            in_quotes = False
        else:
            raise ValueError('CSV 따옴표가 닫히지 않았습니다')
    if row_has_content or row or cell:
        finish_row()
    if not records:
        raise ValueError('데이터 파일에 행이 없습니다')
    return records[0], records[1:]


def _iter_bounded_json_characters(data: bytes):
    """UTF-8 JSON을 청크별 해독해 전체 문서를 문자열로 만들지 않는다."""
    decoder = codecs.getincrementaldecoder('utf-8-sig')(errors='strict')
    decoded_chars = 0
    for offset in range(0, len(data), _DECODE_CHUNK_BYTES):
        decoded = decoder.decode(data[offset:offset + _DECODE_CHUNK_BYTES], final=False)
        decoded_chars += len(decoded)
        if decoded_chars > MAX_CHART_DECODED_CHARS:
            _raise_limit('JSON 해독 문자 수', MAX_CHART_DECODED_CHARS)
        yield from decoded
    decoded = decoder.decode(b'', final=True)
    decoded_chars += len(decoded)
    if decoded_chars > MAX_CHART_DECODED_CHARS:
        _raise_limit('JSON 해독 문자 수', MAX_CHART_DECODED_CHARS)
    yield from decoded


class _JsonCharacterStream:
    """청크 해독된 JSON 문자에서 필요한 부분만 소비한다."""

    def __init__(self, data: bytes) -> None:
        self._characters = iter(_iter_bounded_json_characters(data))

    def take(self) -> str | None:
        return next(self._characters, None)


def _next_json_non_whitespace(stream: _JsonCharacterStream) -> str | None:
    while (character := stream.take()) is not None:
        if not character.isspace():
            return character
    return None


class _JsonTokenBuffer:
    """Accumulate one bounded JSON token in fixed-size string chunks."""

    def __init__(self, *, maximum: int, label: str) -> None:
        self._maximum = maximum
        self._label = label
        self._chunks: list[str] = []
        self._chunk = StringIO()
        self._chunk_chars = 0
        self._length = 0

    def append(self, value: str) -> None:
        self._length += len(value)
        if self._length > self._maximum:
            _raise_limit(self._label, self._maximum)

        remaining_value = value
        while remaining_value:
            remaining_chunk = _JSON_TOKEN_CHUNK_CHARS - self._chunk_chars
            piece = remaining_value[:remaining_chunk]
            self._chunk.write(piece)
            self._chunk_chars += len(piece)
            remaining_value = remaining_value[len(piece):]
            if self._chunk_chars == _JSON_TOKEN_CHUNK_CHARS:
                self._chunks.append(self._chunk.getvalue())
                self._chunk = StringIO()
                self._chunk_chars = 0

    def build(self) -> str:
        if self._chunk_chars:
            self._chunks.append(self._chunk.getvalue())
            self._chunk = StringIO()
            self._chunk_chars = 0
        return ''.join(self._chunks)


def _consume_json_string(stream: _JsonCharacterStream, token: _JsonTokenBuffer) -> None:
    """Append a quoted JSON string already opened by the caller."""
    token.append('"')
    string_chars = 0
    escaped = False
    while (character := stream.take()) is not None:
        token.append(character)
        if escaped:
            string_chars += 1
            escaped = False
        elif character == '\\':
            string_chars += 1
            escaped = True
        elif character == '"':
            return
        else:
            string_chars += 1
        if string_chars > MAX_CHART_DATA_CELL_CHARS:
            _raise_limit('JSON 문자열 문자 수', MAX_CHART_DATA_CELL_CHARS)
    raise ValueError('JSON 문자열이 닫히지 않았습니다')


def _collect_json_string(stream: _JsonCharacterStream) -> str:
    """Collect a bounded standalone JSON string without per-character lists."""
    token = _JsonTokenBuffer(
        maximum=MAX_CHART_DATA_CELL_CHARS + 2,
        label='JSON 문자열 문자 수',
    )
    _consume_json_string(stream, token)
    return token.build()


def _consume_json_scalar_value(
    stream: _JsonCharacterStream,
    token: _JsonTokenBuffer,
    opening: str,
) -> str | None:
    """Append a scalar token and return its object-member delimiter."""
    token.append(opening)
    scalar_chars = 1
    if scalar_chars > MAX_CHART_JSON_SCALAR_TOKEN_CHARS:
        _raise_limit('JSON 스칼라 토큰 문자 수', MAX_CHART_JSON_SCALAR_TOKEN_CHARS)

    while (character := stream.take()) is not None:
        if character in ',}':
            return character
        if character in '{[':
            raise ValueError('JSON 셀은 문자열, 숫자, 불리언 또는 null 이어야 합니다')
        token.append(character)
        if not character.isspace():
            scalar_chars += 1
            if scalar_chars > MAX_CHART_JSON_SCALAR_TOKEN_CHARS:
                _raise_limit('JSON 스칼라 토큰 문자 수', MAX_CHART_JSON_SCALAR_TOKEN_CHARS)
    return None


def _collect_json_object(stream: _JsonCharacterStream) -> str:
    """Preflight one scalar-only object before its bounded JSON decode."""
    token = _JsonTokenBuffer(
        maximum=MAX_CHART_JSON_OBJECT_CHARS,
        label='JSON 객체 문자 수',
    )
    token.append('{')
    member_count = 0
    character = _next_json_non_whitespace(stream)
    if character == '}':
        token.append('}')
        return token.build()

    while True:
        if character != '"':
            raise ValueError('JSON 객체의 키가 올바르지 않습니다')
        member_count += 1
        if member_count > MAX_CHART_DATA_COLUMNS:
            _raise_limit('데이터 열 수', MAX_CHART_DATA_COLUMNS)
        _consume_json_string(stream, token)

        if _next_json_non_whitespace(stream) != ':':
            raise ValueError('JSON 객체 구분자가 올바르지 않습니다')
        token.append(':')
        value_opening = _next_json_non_whitespace(stream)
        if value_opening is None:
            raise ValueError('JSON 객체가 닫히지 않았습니다')
        if value_opening in '{[':
            raise ValueError('JSON 셀은 문자열, 숫자, 불리언 또는 null 이어야 합니다')
        if value_opening == '"':
            _consume_json_string(stream, token)
            separator = _next_json_non_whitespace(stream)
        else:
            separator = _consume_json_scalar_value(stream, token, value_opening)

        if separator == '}':
            token.append('}')
            return token.build()
        if separator != ',':
            raise ValueError('JSON 객체 구분자가 올바르지 않습니다')
        token.append(',')
        character = _next_json_non_whitespace(stream)
        if character is None:
            raise ValueError('JSON 객체가 닫히지 않았습니다')


def _reject_duplicate_json_members(members: list[tuple[str, Any]]) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for name, value in members:
        if name in record:
            raise ValueError('JSON 객체에 중복 키가 있습니다')
        record[name] = value
    return record


def _reject_nonstandard_json_constant(_: str) -> Any:
    raise ValueError('JSON 비표준 숫자 상수는 허용되지 않습니다')


def _parse_finite_json_float(value: str) -> float:
    parsed = float(value)
    if not math.isfinite(parsed):
        _reject_nonstandard_json_constant(value)
    return parsed


def _decode_json_token(token: str) -> Any:
    """Pass only pre-bounded object/string tokens to a strict ``JSONDecoder``."""
    try:
        value, end = json.JSONDecoder(
            object_pairs_hook=_reject_duplicate_json_members,
            parse_constant=_reject_nonstandard_json_constant,
            parse_float=_parse_finite_json_float,
        ).raw_decode(token)
    except (json.JSONDecodeError, OverflowError) as exc:
        raise ValueError('유효하지 않은 JSON 데이터입니다') from exc
    if end != len(token):
        raise ValueError('유효하지 않은 JSON 데이터입니다')
    return value


def _append_json_record(
    token: str,
    *,
    records: list[dict[str, Any]],
    columns: list[str],
    seen_columns: set[str],
) -> None:
    record = _decode_json_token(token)
    if not isinstance(record, dict):
        raise ValueError('JSON 배열의 각 행은 객체여야 합니다')
    if len(record) > MAX_CHART_DATA_COLUMNS:
        _raise_limit('데이터 열 수', MAX_CHART_DATA_COLUMNS)

    for value in record.values():
        if not isinstance(value, (str, int, float, bool, type(None))):
            raise ValueError('JSON 셀은 문자열, 숫자, 불리언 또는 null 이어야 합니다')
        if isinstance(value, str) and len(value) > MAX_CHART_DATA_CELL_CHARS:
            _raise_limit('JSON 셀 문자 수', MAX_CHART_DATA_CELL_CHARS)

    new_columns = [name for name in record if name not in seen_columns]
    if len(columns) + len(new_columns) > MAX_CHART_DATA_COLUMNS:
        _raise_limit('데이터 열 수', MAX_CHART_DATA_COLUMNS)
    _require_rectangular_cell_budget(len(records) + 1, len(columns) + len(new_columns))

    records.append(record)
    columns.extend(new_columns)
    seen_columns.update(new_columns)


def _parse_json_record_array(
    stream: _JsonCharacterStream,
    *,
    max_rows: int,
    records: list[dict[str, Any]],
    columns: list[str],
    seen_columns: set[str],
) -> None:
    character = _next_json_non_whitespace(stream)
    if character == ']':
        return
    while True:
        if character != '{':
            raise ValueError('JSON 배열의 각 행은 객체여야 합니다')
        if len(records) >= max_rows:
            _raise_limit('데이터 행 수', max_rows)
        _append_json_record(
            _collect_json_object(stream),
            records=records,
            columns=columns,
            seen_columns=seen_columns,
        )
        separator = _next_json_non_whitespace(stream)
        if separator == ']':
            return
        if separator != ',':
            raise ValueError('JSON 배열 구분자가 올바르지 않습니다')
        character = _next_json_non_whitespace(stream)
        if character is None:
            raise ValueError('JSON 배열이 닫히지 않았습니다')


def _parse_bounded_json_records(data: bytes, max_rows: int) -> tuple[list[str], list[dict[str, Any]]]:
    """객체 배열 또는 ``{"data": [...]}``만 행별로 검증해 materialize한다."""
    stream = _JsonCharacterStream(data)
    records: list[dict[str, Any]] = []
    columns: list[str] = []
    seen_columns: set[str] = set()
    opening = _next_json_non_whitespace(stream)

    if opening == '[':
        _parse_json_record_array(
            stream,
            max_rows=max_rows,
            records=records,
            columns=columns,
            seen_columns=seen_columns,
        )
    elif opening == '{':
        key_opening = _next_json_non_whitespace(stream)
        if key_opening != '"':
            raise ValueError('JSON 은 객체 배열이거나 data 배열을 가진 객체여야 합니다')
        key = _decode_json_token(_collect_json_string(stream))
        if key != 'data' or _next_json_non_whitespace(stream) != ':':
            raise ValueError('JSON 은 객체 배열이거나 data 배열을 가진 객체여야 합니다')
        if _next_json_non_whitespace(stream) != '[':
            raise ValueError('JSON 은 객체 배열이거나 data 배열을 가진 객체여야 합니다')
        _parse_json_record_array(
            stream,
            max_rows=max_rows,
            records=records,
            columns=columns,
            seen_columns=seen_columns,
        )
        if _next_json_non_whitespace(stream) != '}':
            raise ValueError('JSON 은 객체 배열이거나 data 배열을 가진 객체여야 합니다')
    else:
        raise ValueError('JSON 은 객체 배열이거나 data 배열을 가진 객체여야 합니다')

    if _next_json_non_whitespace(stream) is not None:
        raise ValueError('JSON 뒤에 허용되지 않는 내용이 있습니다')
    if not records:
        raise ValueError('데이터 파일에 행이 없습니다')
    return columns, records


def _read_json(data: bytes, max_rows: int) -> pd.DataFrame:
    try:
        columns, records = _parse_bounded_json_records(data, max_rows)
    except (UnicodeDecodeError, json.JSONDecodeError, RecursionError, ValueError) as exc:
        raise ValueError(f'유효하지 않은 JSON 데이터입니다: {exc}') from exc

    _require_rectangular_cell_budget(len(records), len(columns))
    rows = [[record.get(name) for name in columns] for record in records]
    return _dataframe_from_rows(columns, rows)


def _canonical_excel_member_name(name: str) -> str:
    path = PurePosixPath(name)
    if (
        not name
        or '\x00' in name
        or '\\' in name
        or name.startswith('/')
        or re.match(r'^[A-Za-z]:', name)
        or '..' in path.parts
        or path.is_absolute()
    ):
        raise ValueError('XLSX 에 안전하지 않은 ZIP 멤버 이름이 있습니다')
    return str(path)


def _excel_xml_local_name(tag: str) -> str:
    return tag.rsplit('}', 1)[-1]


def _iter_bounded_excel_xml_events(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    max_bytes: int | None = None,
    limit_label: str = 'XLSX 메타데이터 바이트',
):
    """Stream XML in fixed chunks and release completed elements immediately."""
    if max_bytes is None:
        max_bytes = MAX_CHART_XLSX_METADATA_BYTES
    if max_bytes < 1:
        raise ValueError('XLSX XML 바이트 상한은 1 이상이어야 합니다')

    parser = XMLPullParser(events=('start', 'end'))
    total = 0
    with archive.open(info) as member:
        while True:
            remaining = max_bytes + 1 - total
            chunk = member.read(min(_DECODE_CHUNK_BYTES, remaining))
            if not chunk:
                break
            total += len(chunk)
            if total > max_bytes:
                _raise_limit(limit_label, max_bytes)
            parser.feed(chunk)
            for event, element in parser.read_events():
                yield event, element
                if event == 'end':
                    element.clear()
    parser.close()
    for event, element in parser.read_events():
        yield event, element
        if event == 'end':
            element.clear()


def _preflight_excel_metadata(archive: zipfile.ZipFile, infos: dict[str, zipfile.ZipInfo]) -> None:
    """openpyxl 전에 공유 문자열·스타일·콘텐츠 형식의 할당량을 검증한다."""
    shared_strings = infos.get('xl/sharedStrings.xml')
    if shared_strings is not None:
        count = 0
        current_chars: int | None = None
        try:
            for event, element in _iter_bounded_excel_xml_events(archive, shared_strings):
                name = _excel_xml_local_name(element.tag)
                if event == 'start' and name == 'si':
                    current_chars = 0
                elif event == 'end' and name == 't' and current_chars is not None:
                    current_chars += len(element.text or '')
                    if current_chars > MAX_CHART_DATA_CELL_CHARS:
                        _raise_limit('XLSX 공유 문자열 문자 수', MAX_CHART_DATA_CELL_CHARS)
                elif event == 'end' and name == 'si':
                    count += 1
                    if count > MAX_CHART_XLSX_SHARED_STRINGS:
                        _raise_limit('XLSX 공유 문자열 수', MAX_CHART_XLSX_SHARED_STRINGS)
                    current_chars = None
        except ParseError as exc:
            raise ValueError('XLSX 공유 문자열 메타데이터가 손상되었습니다') from exc

    styles = infos.get('xl/styles.xml')
    if styles is not None:
        style_records = 0
        number_formats = 0
        specific_counts: dict[str, int] = {}
        specific_limits = {
            'font': ('XLSX 글꼴 스타일 수', MAX_CHART_XLSX_FONTS),
            'fill': ('XLSX 채우기 스타일 수', MAX_CHART_XLSX_FILLS),
            'border': ('XLSX 테두리 스타일 수', MAX_CHART_XLSX_BORDERS),
            'alignment': ('XLSX 맞춤 스타일 수', MAX_CHART_XLSX_ALIGNMENTS),
            'protection': ('XLSX 보호 스타일 수', MAX_CHART_XLSX_PROTECTIONS),
            'cellStyle': ('XLSX 명명 스타일 수', MAX_CHART_XLSX_NAMED_STYLES),
            'dxf': ('XLSX 차등 스타일 수', MAX_CHART_XLSX_DIFFERENTIAL_STYLES),
            'stop': ('XLSX 그라데이션 중지점 수', MAX_CHART_XLSX_GRADIENT_STOPS),
            'tableStyleElement': (
                'XLSX 표 스타일 요소 수',
                MAX_CHART_XLSX_TABLE_STYLE_ELEMENTS,
            ),
            'indexedColor': ('XLSX 색인 색상 수', MAX_CHART_XLSX_INDEXED_COLORS),
            'mruColor': ('XLSX 최근 색상 수', MAX_CHART_XLSX_MRU_COLORS),
        }
        allocation_tags = _XLSX_STYLE_ALLOCATION_TAGS
        try:
            for event, element in _iter_bounded_excel_xml_events(archive, styles):
                if event != 'end':
                    continue
                name = _excel_xml_local_name(element.tag)
                if name in allocation_tags:
                    style_records += 1
                    if style_records > MAX_CHART_XLSX_STYLE_RECORDS:
                        _raise_limit('XLSX 스타일 수', MAX_CHART_XLSX_STYLE_RECORDS)
                if name == 'numFmt':
                    number_formats += 1
                    if number_formats > MAX_CHART_XLSX_NUMBER_FORMATS:
                        _raise_limit('XLSX 사용자 지정 숫자 형식 수', MAX_CHART_XLSX_NUMBER_FORMATS)
                elif name in specific_limits:
                    label, limit = specific_limits[name]
                    count = specific_counts.get(name, 0) + 1
                    specific_counts[name] = count
                    if count > limit:
                        _raise_limit(label, limit)
        except ParseError as exc:
            raise ValueError('XLSX 스타일 메타데이터가 손상되었습니다') from exc

    content_types = infos.get('[Content_Types].xml')
    if content_types is not None:
        count = 0
        try:
            for event, element in _iter_bounded_excel_xml_events(archive, content_types):
                if event == 'end' and _excel_xml_local_name(element.tag) in {'Default', 'Override'}:
                    count += 1
                    if count > MAX_CHART_XLSX_CONTENT_TYPES:
                        _raise_limit('XLSX 콘텐츠 형식 수', MAX_CHART_XLSX_CONTENT_TYPES)
        except ParseError as exc:
            raise ValueError('XLSX 콘텐츠 형식 메타데이터가 손상되었습니다') from exc


@dataclass(frozen=True)
class _ExcelWorksheetBounds:
    max_row: int
    max_column: int
    worksheet_index: int


def _excel_xml_tag(element, namespace: str, name: str) -> bool:
    return element.tag == f'{{{namespace}}}{name}'


def _excel_xml_attribute(element, name: str) -> str | None:
    return element.attrib.get(name)


def _parse_bounded_excel_row(value: str | None, *, maximum: int) -> int:
    if not value or any(character < '0' or character > '9' for character in value):
        raise ValueError('XLSX 워크시트 행 좌표가 올바르지 않습니다')
    normalized = value.lstrip('0')
    if not normalized:
        raise ValueError('XLSX 워크시트 행 좌표가 올바르지 않습니다')
    maximum_text = str(maximum)
    if len(normalized) > len(maximum_text) or (
        len(normalized) == len(maximum_text) and normalized > maximum_text
    ):
        _raise_limit('XLSX 워크시트 행 좌표', maximum)
    return int(normalized)


def _parse_excel_cell_reference(value: str | None, *, max_row: int) -> tuple[int, int]:
    match = _XLSX_CELL_REFERENCE_RE.fullmatch(value or '')
    if match is None:
        raise ValueError('XLSX 워크시트 셀 좌표가 올바르지 않습니다')

    column = 0
    for character in match.group(1):
        column = column * 26 + ord(character) - ord('A') + 1
    if column > MAX_CHART_DATA_COLUMNS:
        _raise_limit('데이터 열 수', MAX_CHART_DATA_COLUMNS)
    return column, _parse_bounded_excel_row(match.group(2), maximum=max_row)


def _parse_excel_dimension_bounds(
    value: str | None,
    *,
    max_row: int,
) -> tuple[int, int, int, int]:
    references = (value or '').split(':')
    if len(references) not in {1, 2}:
        raise ValueError('XLSX 워크시트 범위가 올바르지 않습니다')
    start_column, start_row = _parse_excel_cell_reference(references[0], max_row=max_row)
    end_column, end_row = _parse_excel_cell_reference(references[-1], max_row=max_row)
    if end_column < start_column or end_row < start_row:
        raise ValueError('XLSX 워크시트 범위가 올바르지 않습니다')
    return start_column, start_row, end_column, end_row




def _first_excel_worksheet_info(
    archive: zipfile.ZipFile,
    infos: dict[str, zipfile.ZipInfo],
) -> tuple[zipfile.ZipInfo, int]:
    """Resolve the first visible worksheet exactly as workbook metadata declares."""
    workbook = infos.get('xl/workbook.xml')
    relationships = infos.get('xl/_rels/workbook.xml.rels')
    if workbook is None or relationships is None:
        raise ValueError('XLSX 워크시트 메타데이터가 없습니다')

    targets: dict[str, zipfile.ZipInfo] = {}
    relationship_ids: set[str] = set()
    try:
        relationship_count = 0
        relationship_stack: list[str] = []
        relationship_root = f'{{{_OOXML_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationships'
        for event, element in _iter_bounded_excel_xml_events(archive, relationships):
            if event == 'start':
                if not relationship_stack and element.tag != relationship_root:
                    raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다')
                relationship_stack.append(element.tag)
                continue
            if not relationship_stack or relationship_stack[-1] != element.tag:
                raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다')
            if (
                len(relationship_stack) == 2
                and relationship_stack[0] == relationship_root
                and _excel_xml_tag(
                    element,
                    _OOXML_PACKAGE_RELATIONSHIPS_NAMESPACE,
                    'Relationship',
                )
            ):
                relationship_count += 1
                if relationship_count > MAX_CHART_XLSX_RELATIONSHIPS:
                    _raise_limit('XLSX 워크시트 관계 수', MAX_CHART_XLSX_RELATIONSHIPS)
                relationship_id = _excel_xml_attribute(element, 'Id')
                if not relationship_id:
                    raise ValueError('XLSX 워크시트 관계 ID가 올바르지 않습니다')
                if relationship_id in relationship_ids:
                    raise ValueError('XLSX 워크시트 관계 ID가 중복되었습니다')
                relationship_ids.add(relationship_id)

                relationship_type = _excel_xml_attribute(element, 'Type')
                if relationship_type == _OOXML_WORKSHEET_RELATIONSHIP_TYPE:
                    target = _excel_xml_attribute(element, 'Target')
                    if not target or (
                        _excel_xml_attribute(element, 'TargetMode') or ''
                    ).lower() == 'external':
                        raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다')
                    resolved = (
                        target.lstrip('/')
                        if target.startswith('/')
                        else str(PurePosixPath('xl') / target)
                    )
                    canonical_target = _canonical_excel_member_name(resolved)
                    target_info = infos.get(canonical_target)
                    if target_info is None or target_info.is_dir():
                        raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다')
                    targets[relationship_id] = target_info
            relationship_stack.pop()
        if relationship_stack:
            raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다')

        workbook_records = 0
        specific_counts: dict[str, int] = {}
        specific_limits = {
            'sheet': ('XLSX 워크시트 수', MAX_CHART_XLSX_WORKBOOK_SHEETS),
            'definedName': ('XLSX 정의 이름 수', MAX_CHART_XLSX_DEFINED_NAMES),
            'externalReference': (
                'XLSX 외부 참조 수',
                MAX_CHART_XLSX_EXTERNAL_REFERENCES,
            ),
            'pivotCache': ('XLSX 피벗 캐시 수', MAX_CHART_XLSX_PIVOT_CACHES),
            'workbookView': ('XLSX 통합 문서 보기 수', MAX_CHART_XLSX_WORKBOOK_VIEWS),
            'functionGroup': ('XLSX 함수 그룹 수', MAX_CHART_XLSX_FUNCTION_GROUPS),
            'customWorkbookView': (
                'XLSX 사용자 통합 문서 보기 수',
                MAX_CHART_XLSX_CUSTOM_WORKBOOK_VIEWS,
            ),
            'fileRecoveryPr': (
                'XLSX 복구 레코드 수',
                MAX_CHART_XLSX_FILE_RECOVERY_RECORDS,
            ),
            'smartTagType': ('XLSX 스마트 태그 유형 수', MAX_CHART_XLSX_SMART_TAG_TYPES),
            'webPublishObject': (
                'XLSX 웹 게시 객체 수',
                MAX_CHART_XLSX_WEB_PUBLISH_OBJECTS,
            ),
        }
        workbook_root = f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}workbook'
        workbook_allocation_parents = {
            'workbookView': f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}bookViews',
            'sheet': f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}sheets',
            'functionGroup': f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}functionGroups',
            'externalReference': f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}externalReferences',
            'definedName': f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}definedNames',
            'customWorkbookView': f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}customWorkbookViews',
            'pivotCache': f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}pivotCaches',
            'smartTagType': f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}smartTagTypes',
            'fileRecoveryPr': workbook_root,
            'webPublishObject': f'{{{_OOXML_SPREADSHEET_NAMESPACE}}}webPublishObjects',
        }
        selected: tuple[zipfile.ZipInfo, int] | None = None
        worksheet_index = 0
        workbook_stack: list[str] = []
        for event, element in _iter_bounded_excel_xml_events(archive, workbook):
            if event == 'start':
                if not workbook_stack and element.tag != workbook_root:
                    raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다')
                workbook_stack.append(element.tag)
                continue
            if not workbook_stack or workbook_stack[-1] != element.tag:
                raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다')
            name = _excel_xml_local_name(element.tag)
            parent = workbook_stack[-2] if len(workbook_stack) > 1 else None
            if (
                _excel_xml_tag(element, _OOXML_SPREADSHEET_NAMESPACE, name)
                and name in _XLSX_WORKBOOK_ALLOCATION_TAGS
                and workbook_allocation_parents[name] == parent
            ):
                workbook_records += 1
                if workbook_records > MAX_CHART_XLSX_WORKBOOK_RECORDS:
                    _raise_limit('XLSX 통합 문서 메타데이터 수', MAX_CHART_XLSX_WORKBOOK_RECORDS)
                if name in specific_limits:
                    label, limit = specific_limits[name]
                    count = specific_counts.get(name, 0) + 1
                    specific_counts[name] = count
                    if count > limit:
                        _raise_limit(label, limit)

                if name == 'sheet':
                    relationship_id = element.attrib.get(_OOXML_RELATIONSHIP_ID_ATTRIBUTE)
                    unqualified_id = _excel_xml_attribute(element, 'id')
                    if relationship_id is not None and unqualified_id is not None:
                        raise ValueError('XLSX 워크시트 관계 ID가 모호합니다')
                    if not relationship_id:
                        raise ValueError('XLSX 워크시트 관계 ID가 올바르지 않습니다')
                    if relationship_id not in relationship_ids:
                        raise ValueError('XLSX 워크시트 관계 ID가 없습니다')
                    target_info = targets.get(relationship_id)
                    if target_info is not None:
                        selected_index = worksheet_index
                        worksheet_index += 1
                        state = _excel_xml_attribute(element, 'state')
                        if selected is None and state in {None, 'visible'}:
                            selected = (target_info, selected_index)
            workbook_stack.pop()
        if workbook_stack:
            raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다')
    except ParseError as exc:
        raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다') from exc
    if selected is None:
        raise ValueError('XLSX 워크시트가 없습니다')
    return selected


def _preflight_selected_excel_worksheet(
    archive: zipfile.ZipFile,
    infos: dict[str, zipfile.ZipInfo],
    *,
    max_rows: int,
) -> _ExcelWorksheetBounds:
    """Validate selected-sheet coordinates and cell nodes before openpyxl."""
    sheet, worksheet_index = _first_excel_worksheet_info(archive, infos)
    maximum_physical_row = max_rows + 1
    dimension_bounds: tuple[int, int, int, int] | None = None
    dimension_column = 0
    dimension_row = 0
    observed_column = 0
    observed_row = 0
    observed_cell_minimum: tuple[int, int] | None = None
    observed_cell_maximum: tuple[int, int] | None = None
    previous_row = 0
    physical_rows = 0
    active_row: int | None = None
    active_cell: tuple[int, int] | None = None
    cell_payload_depth = 0
    direct_cell_payload_kinds: set[str] = set()
    previous_column = 0
    cells_in_active_row = 0
    cell_nodes = 0
    inline_payloads = 0
    inline_payload_chars = 0

    def require_dimension_contains(column: int, row: int) -> None:
        if dimension_bounds is None:
            return
        start_column, start_row, end_column, end_row = dimension_bounds
        if not (start_column <= column <= end_column and start_row <= row <= end_row):
            raise ValueError('XLSX 워크시트 범위와 셀 좌표가 일치하지 않습니다')

    try:
        for event, element in _iter_bounded_excel_xml_events(
            archive,
            sheet,
            max_bytes=MAX_CHART_XLSX_MEMBER_BYTES,
            limit_label='XLSX 워크시트 XML 바이트',
        ):
            name = _excel_xml_local_name(element.tag)
            if not _excel_xml_tag(element, _OOXML_SPREADSHEET_NAMESPACE, name):
                if active_cell is not None:
                    if event == 'start':
                        cell_payload_depth += 1
                    elif cell_payload_depth < 1:
                        raise ValueError('XLSX 워크시트 셀 연결이 올바르지 않습니다')
                    else:
                        cell_payload_depth -= 1
                continue
            if event == 'start':
                if name == 'dimension':
                    if dimension_bounds is not None:
                        raise ValueError('XLSX 워크시트 범위가 중복되었습니다')
                    dimension_bounds = _parse_excel_dimension_bounds(
                        _excel_xml_attribute(element, 'ref'),
                        max_row=maximum_physical_row,
                    )
                    _, _, dimension_column, dimension_row = dimension_bounds
                    if observed_cell_minimum is not None and observed_cell_maximum is not None:
                        require_dimension_contains(*observed_cell_minimum)
                        require_dimension_contains(*observed_cell_maximum)
                elif name == 'row':
                    if active_row is not None:
                        raise ValueError('XLSX 워크시트 행 연결이 올바르지 않습니다')
                    row = _parse_bounded_excel_row(
                        _excel_xml_attribute(element, 'r'),
                        maximum=maximum_physical_row,
                    )
                    if row <= previous_row:
                        raise ValueError('XLSX 워크시트 행 좌표가 정렬되지 않았습니다')
                    gap = row - previous_row - 1
                    if gap > max_rows:
                        _raise_limit('XLSX 빈 행 간격', max_rows)
                    previous_row = row
                    observed_row = max(observed_row, row)
                    physical_rows += 1
                    if physical_rows > maximum_physical_row:
                        _raise_limit('XLSX 물리 행 수', maximum_physical_row)
                    active_row = row
                    active_cell = None
                    previous_column = 0
                    cells_in_active_row = 0
                elif name == 'c':
                    if active_row is None or active_cell is not None:
                        raise ValueError('XLSX 워크시트 셀 행 연결이 올바르지 않습니다')
                    column, row = _parse_excel_cell_reference(
                        _excel_xml_attribute(element, 'r'),
                        max_row=maximum_physical_row,
                    )
                    if row != active_row:
                        raise ValueError('XLSX 워크시트 셀 행 연결이 올바르지 않습니다')
                    if column <= previous_column:
                        raise ValueError('XLSX 워크시트 셀 좌표가 정렬되지 않았습니다')
                    require_dimension_contains(column, row)
                    previous_column = column
                    cells_in_active_row += 1
                    if cells_in_active_row > MAX_CHART_DATA_COLUMNS:
                        _raise_limit('XLSX 행별 셀 수', MAX_CHART_DATA_COLUMNS)
                    cell_nodes += 1
                    if cell_nodes > MAX_CHART_DATA_CELLS:
                        _raise_limit('데이터 셀 수', MAX_CHART_DATA_CELLS)
                    observed_column = max(observed_column, column)
                    observed_row = max(observed_row, row)
                    if observed_cell_minimum is None:
                        observed_cell_minimum = (column, row)
                        observed_cell_maximum = (column, row)
                    else:
                        observed_cell_minimum = (
                            min(observed_cell_minimum[0], column),
                            min(observed_cell_minimum[1], row),
                        )
                        observed_cell_maximum = (
                            max(observed_cell_maximum[0], column),
                            max(observed_cell_maximum[1], row),
                        )
                    active_cell = (column, row)
                    inline_payloads = 0
                    inline_payload_chars = 0
                    cell_payload_depth = 0
                    direct_cell_payload_kinds = set()
                elif active_cell is not None:
                    if cell_payload_depth == 0 and name in {'f', 'is', 'v'}:
                        if name in direct_cell_payload_kinds:
                            raise ValueError(f'XLSX 셀 {name} 페이로드가 중복되었습니다')
                        direct_cell_payload_kinds.add(name)
                    if name in {'f', 'is', 'r', 'rPh', 't', 'v'}:
                        inline_payloads += 1
                        if inline_payloads > MAX_CHART_XLSX_INLINE_PAYLOADS_PER_CELL:
                            _raise_limit(
                                'XLSX 인라인 셀 페이로드 수',
                                MAX_CHART_XLSX_INLINE_PAYLOADS_PER_CELL,
                            )
                    cell_payload_depth += 1
            elif event == 'end':
                if name == 'c':
                    if active_cell is None or cell_payload_depth != 0:
                        raise ValueError('XLSX 워크시트 셀 연결이 올바르지 않습니다')
                    active_cell = None
                elif name == 'row':
                    if active_row is None or active_cell is not None:
                        raise ValueError('XLSX 워크시트 행 연결이 올바르지 않습니다')
                    active_row = None
                elif active_cell is not None:
                    if cell_payload_depth < 1:
                        raise ValueError('XLSX 워크시트 셀 연결이 올바르지 않습니다')
                    if name in {'f', 't', 'v'}:
                        inline_payload_chars += len(element.text or '')
                        if inline_payload_chars > MAX_CHART_DATA_CELL_CHARS:
                            _raise_limit('XLSX 셀 문자 수', MAX_CHART_DATA_CELL_CHARS)
                    cell_payload_depth -= 1
    except ParseError as exc:
        raise ValueError('XLSX 워크시트 XML이 손상되었습니다') from exc

    if active_row is not None or active_cell is not None:
        raise ValueError('XLSX 워크시트 행 연결이 올바르지 않습니다')
    maximum_row = max(dimension_row, observed_row, 1)
    maximum_column = max(dimension_column, observed_column, 1)
    physical_data_rows = maximum_row - 1
    if physical_data_rows > max_rows:
        _raise_limit('데이터 행 수', max_rows)
    _require_rectangular_cell_budget(physical_data_rows, maximum_column)
    return _ExcelWorksheetBounds(
        max_row=maximum_row,
        max_column=maximum_column,
        worksheet_index=worksheet_index,
    )

def _preflight_excel_archive(data: bytes, max_rows: int) -> _ExcelWorksheetBounds:
    """EOCD/ZIP64와 OOXML 메타데이터를 검증한 뒤에만 ZipFile을 연다."""
    if len(data) > MAX_CHART_XLSX_COMPRESSED_BYTES:
        raise UploadSizeLimitExceeded('ZIP 압축 업로드 총 용량 상한을 초과했습니다')
    source = BytesIO(data)
    try:
        preflight = preflight_zip_central_directory(
            source,
            max_files=MAX_CHART_XLSX_ENTRIES,
            max_compressed_bytes=MAX_CHART_XLSX_COMPRESSED_BYTES,
            max_filename_bytes=MAX_CHART_XLSX_FILENAME_BYTES,
            max_central_directory_bytes=MAX_CHART_XLSX_CENTRAL_DIRECTORY_BYTES,
            canonicalize_filename=_canonical_excel_member_name,
            filename_limit_message='XLSX ZIP 멤버 이름 바이트 상한을 초과했습니다',
        )
        names: set[str] = set()
        for member in preflight.members:
            if member.canonical_name in names:
                raise ValueError('XLSX ZIP 에 중복 멤버 이름이 있습니다')
            names.add(member.canonical_name)

        with zipfile.ZipFile(source) as archive:
            infos = archive.infolist()
            if len(infos) != len(preflight.members):
                raise ValueError('XLSX 파일이 손상되었거나 읽을 수 없습니다')

            metadata_infos: dict[str, zipfile.ZipInfo] = {}
            expanded_total = 0
            for info, member in zip(infos, preflight.members, strict=True):
                if info.filename != member.filename or info.is_dir() != member.is_directory:
                    raise ValueError('XLSX 파일이 손상되었거나 읽을 수 없습니다')
                if info.is_dir():
                    continue
                if info.file_size > MAX_CHART_XLSX_MEMBER_BYTES:
                    _raise_limit('XLSX ZIP 멤버 확장 바이트', MAX_CHART_XLSX_MEMBER_BYTES)
                expanded_total += info.file_size
                if expanded_total > MAX_CHART_XLSX_EXPANDED_BYTES:
                    _raise_limit('XLSX ZIP 총 확장 바이트', MAX_CHART_XLSX_EXPANDED_BYTES)
                if info.file_size and (
                    not info.compress_size
                    or info.file_size > info.compress_size * MAX_CHART_XLSX_COMPRESSION_RATIO
                ):
                    _raise_limit('XLSX ZIP 압축률', MAX_CHART_XLSX_COMPRESSION_RATIO)
                metadata_infos[info.filename] = info
            _preflight_excel_metadata(archive, metadata_infos)
            return _preflight_selected_excel_worksheet(
                archive,
                metadata_infos,
                max_rows=max_rows,
            )
    except (zipfile.BadZipFile, EOFError, NotImplementedError, OSError, RuntimeError, zlib.error) as exc:
        raise ValueError('XLSX 파일이 손상되었거나 읽을 수 없습니다') from exc


def _read_excel(data: bytes, max_rows: int) -> pd.DataFrame:
    worksheet_bounds = _preflight_excel_archive(data, max_rows)
    if find_spec('openpyxl') is None:
        raise ValueError('xlsx 지원 라이브러리(openpyxl)가 설치되어 있지 않습니다. CSV 나 JSON 을 사용하세요.')

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    workbook = None
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True, keep_links=False)
        if not workbook.worksheets:
            raise ValueError('XLSX 워크시트가 없습니다')
        if len(workbook.worksheets) <= worksheet_bounds.worksheet_index:
            raise ValueError('XLSX 워크시트 메타데이터가 손상되었습니다')
        sheet = workbook.worksheets[worksheet_bounds.worksheet_index]
        rows = sheet.iter_rows(
            values_only=True,
            min_row=1,
            max_row=worksheet_bounds.max_row,
            min_col=1,
            max_col=worksheet_bounds.max_column,
        )
        try:
            header = list(next(rows))
        except StopIteration as exc:
            raise ValueError('데이터 파일에 행이 없습니다') from exc
        if len(header) > MAX_CHART_DATA_COLUMNS:
            _raise_limit('데이터 열 수', MAX_CHART_DATA_COLUMNS)
        for value in header:
            if isinstance(value, str) and len(value) > MAX_CHART_DATA_CELL_CHARS:
                _raise_limit('XLSX 셀 문자 수', MAX_CHART_DATA_CELL_CHARS)

        records: list[list[Any]] = []
        for values in rows:
            row = list(values)
            if not any(value is not None for value in row):
                continue
            if len(records) >= max_rows:
                _raise_limit('데이터 행 수', max_rows)
            _require_rectangular_cell_budget(len(records) + 1, len(header))
            for value in row:
                if isinstance(value, str) and len(value) > MAX_CHART_DATA_CELL_CHARS:
                    _raise_limit('XLSX 셀 문자 수', MAX_CHART_DATA_CELL_CHARS)
            records.append(row)
        return _dataframe_from_rows(header, records)
    except (
        InvalidFileException,
        KeyError,
        ParseError,
        zipfile.BadZipFile,
        EOFError,
        NotImplementedError,
        OSError,
        RuntimeError,
        zlib.error,
    ) as exc:
        raise ValueError('XLSX 파일이 손상되었거나 읽을 수 없습니다') from exc
    finally:
        if workbook is not None:
            workbook.close()


def load_dataframe(filename: str, data: bytes, max_rows: int) -> pd.DataFrame:
    """업로드를 bounded parser로 로드하고 행수 상한/빈 파일을 시스템 경계에서 막는다."""
    if max_rows < 1:
        raise ValueError('데이터 행 상한은 1 이상이어야 합니다')
    suffix = Path(filename or '').suffix.lower()
    if suffix not in SUPPORTED_DATA_SUFFIXES:
        raise ValueError(f'지원하지 않는 데이터 확장자입니다: {suffix or "(없음)"}')
    if suffix == '.csv':
        frame = _read_csv(data, max_rows)
    elif suffix in _EXCEL_SUFFIXES:
        frame = _read_excel(data, max_rows)
    else:
        frame = _read_json(data, max_rows)

    if frame.empty:
        raise ValueError('데이터 파일에 행이 없습니다')
    if len(frame) > max_rows:
        _raise_limit('데이터 행 수', max_rows)
    frame = frame.copy()
    frame.columns = _unique_columns(frame.columns)
    return frame.replace([np.inf, -np.inf], np.nan)


def dataframe_profile(frame: pd.DataFrame, sample_rows: int = 8) -> dict[str, Any]:
    """열별 dtype/결측/유일값과 앞부분 샘플을 JSON-safe 하게 요약한다."""
    columns = []
    for name in frame.columns:
        series = frame[name]
        columns.append(
            {
                'name': str(name),
                'dtype': str(series.dtype),
                'non_null': int(series.notna().sum()),
                'null': int(series.isna().sum()),
                'unique': int(series.nunique(dropna=True)),
                'numeric': bool(pd.api.types.is_numeric_dtype(series)),
                'datetime': bool(pd.api.types.is_datetime64_any_dtype(series)),
            }
        )
    head = frame.head(sample_rows)
    sample = head.where(pd.notna(head), None).to_dict(orient='records')
    # Timestamp/numpy 스칼라를 문자열로 낮춰 JSON 직렬화 안전을 보장한다.
    safe_sample = json.loads(json.dumps(sample, default=str, ensure_ascii=False))
    return {
        'row_count': int(len(frame)),
        'column_count': int(len(frame.columns)),
        'columns': columns,
        'sample': safe_sample,
    }
